import math
import time
from openpyxl import load_workbook
#from openpyxl import Workbook
# import xlwt
# import pandas as pd


#def writeListInXlsx(tableName, rows):
#    book = Workbook()
#    sheet = book.active
#    for row in rows:
#        # print(row)
#        sheet.append(row)
#
#    book.save('{}.xlsx'.format(tableName))

# Загрузка матрицы сопряженности линий из excel-файла
def loadMatrix(wbName, sheetName, cellsInd):
    # Load in the workbook
    wb = load_workbook(wbName)
    # Get a sheet by name
    sheet = wb.get_sheet_by_name(sheetName)

    listMatrix = []
    arrayRowLine = []
    for cellObj in sheet[cellsInd[0]:cellsInd[1]]:
        for cells in cellObj:
            arrayRowLine.append(cells.value)
        listMatrix.append(arrayRowLine.copy())
        arrayRowLine.clear()
    return listMatrix


# Загрузка данных по ТП в словарь из excel-файла в виде:
  # №_линии: (№_ТП, номин_мощн, загрузка)
def loadDataTP(wbName, sheetName, cellsInd):
    wb = load_workbook(wbName)
    sheet = wb.get_sheet_by_name(sheetName)

    dictPowerLine = {}
    for cellObj in sheet[cellsInd[0]:cellsInd[1]]:
        dictPowerLine[int(cellObj[0].value)] = (str(cellObj[1].value),
                                                int(cellObj[2].value),
                                                float(cellObj[3].value),
                                                int(cellObj[4].value))
    return dictPowerLine


# Загрузка объединенных позиций
def loadUnionPosit(wbName, sheetName, cellsInd):
    wb = load_workbook(wbName)
    sheet = wb.get_sheet_by_name(sheetName)

    listPosit = []

    for cellObj in sheet[cellsInd[0]:cellsInd[1]]:
        listPosit.append((str(cellObj[0].value)).split())

    listPosit.append([])

    return listPosit


 # Для каждой позиции расчитывается количество следующих за ней линий
def protectLines(inMatrix, numbPillar):
    dictProtectLines = {}
    for i in range(len(inMatrix)):
        #dictProtectLines[i+1] = sum(inMatrix[i]) #np.array(sum(inMatrix[i]))
        dictProtectLines[i+1] = sum([x * y for x, y in zip(inMatrix[i], numbPillar)])
    return dictProtectLines


# защищаемая мощность для каждой позиции КА
def functProtectConsumer(dictProtectPower, matrixGraph, amountLines):
    arrayProtectPower = []
    for posKA1 in range(0, amountLines):
        protectPower = 0
        for posKA2 in range(0, amountLines):
            if matrixGraph[posKA1][posKA2] == 1:
                listPower = dictProtectPower[posKA2 + 1]
                protectPower += listPower[1] * listPower[2]
        arrayProtectPower.append(protectPower)
    return arrayProtectPower


# проверка на наличие соседних позиций.
def notIsNeigh(listPos, dictNeigh):
    breakGenPos = True
    rowEl = listPos[-1]-1
    for el in reversed(listPos[:-1]):
        if dictNeigh[rowEl][el-1] == 1:
            breakGenPos = False
            break
    return breakGenPos


# поиск независимых позиций
def searchIndepPos(positions, matrixGraph):
    lstIndependPositions = []
    lenPos = len(positions)
    for posI in range(lenPos-1):
        isBreak = False
        for posJ in range(posI+1,lenPos):
            if matrixGraph[positions[posJ]-1][positions[posI]-1] == 1:
                isBreak = True
                break
        if isBreak: continue
        lstIndependPositions.append(positions[posI])
    lstIndependPositions.append(positions[-1])
    return tuple(lstIndependPositions)


# расчет кол-ва защищаемых линий
def calculProtLines(positions, matrixGraph, dictProtLines):
    lstProtLines = []
    dictPosAndPrL = {}
    for iPos in positions:

        positProtLines = dictProtLines[iPos]

        for jPos in list(dictPosAndPrL.keys()):

            if matrixGraph[iPos - 1][jPos - 1] == 1:
                positProtLines -= dictProtLines[jPos] #
                del dictPosAndPrL[jPos]
       
        lstProtLines.append(positProtLines)
        dictPosAndPrL[iPos] = positProtLines

    return lstProtLines

#-----------------------------------------------------------------------------------
#--------------------------------------- РАССЧЕТНАЯ ЧАСТЬ --------------------------
#-----------------------------------------------------------------------------------

feederName = input("Введите название фидера (имя папки): ")
wbName = 'C:/EnerTechCenter/ДанныеПоСетям/{}/'.format(feederName)

with open(wbName+"cellsInd.txt") as file_cells:
    for i in range(2):
        if i==0: cellDataTP = tuple(file_cells.readline().split())
        else: cellMatrix = tuple(file_cells.readline().split())

# Загрузка словаря из эксель с данными о ТП (позиция: №, ном.мощн, ст.загрузки)
dictDataTP = loadDataTP(wbName=wbName+'dataTP.xlsx', sheetName='dataTP', cellsInd=cellDataTP)


# Загрузка матрицы зависимости позиций
matrixGraph = loadMatrix(wbName=wbName+'dependMatr.xlsx', sheetName='Sheet1', cellsInd=cellMatrix)

# Формирование словаря, содержащего для каждой позиции КА количесвто зависимых позиций
dictProtectLines = protectLines(matrixGraph, [tpData[3] for tpData in dictDataTP.values()])

# Получение количества позиций на данной сети
AMOUNT_LINES = len(dictDataTP)
AMOUNT_PILLAR = dictProtectLines[1]

# Расчет защищаемой мощности каждой позицией при условии, что нет ни единой КА
arrayProtectPower = functProtectConsumer(dictDataTP, matrixGraph, AMOUNT_LINES) # зщищаемая млщность каждой КА

#relevantPositions = {'к8'  : [8, 133],  'к20' : [20, 102], 'к21' : [21, 97], 'к35' : [35, 83],
#                  'к36' : [36, 82],  'к37' : [37, 84],  'к41' : [41, 87], 'к42' : [42, 88], 
#                  'к48' : [48, 94], 'к60' : [60, 106],  'к68' : [72, 126], 'к69' : [73, 125], 
#                  'п6'  : [146, 6], 'п8'  : [144, 8],   'п15' : [141, 15], 'п37' : [93, 37],  
#                  'п42' : [81, 42], 'п43' : [88, 43],   'р6'  : [67, 62]}

relevantPositions = {'б_56' : [65, 152], 'б_52' : [59, 149], 'б_37' : [40, 139], 'б_35' : [36, 98],
                     'б_34' : [35, 97],  'б_28' : [28, 91],  'б_27' : [27, 127], 'б_25' : [25, 92],  
                     'б_19' : [19, 117], 'б_18' : [18, 108], 'б_17' : [17, 118], 'р6'   : [29, 84],
                     'г_73' : [107, 73], 'г_62' : [95, 62],  'г_48' : [85, 48],  'г_44' : [75, 44],
                     'г_43' : [80, 43],  'г_36' : [87, 36],  'г_35' : [93, 35],  'г_34' : [92, 34], 
                     'г_20' : [110, 20], 'г_18' : [125, 18], 'г_13' : [146, 13], 'г_8'  : [159, 8]}

#currentCombPos = [['к69', 'к68', 'к42', 'к41', 'к36', 'к21', 'к20', 'к8' , 'п43', 'п42', 'п37', 'п15', 'п8', 'п6'],
#           ['к69', 'к68', 'к42', 'к41', 'к36', 'к21', 'к20', 'к8' , 'п43', 'п42', 'п37', 'п15', 'п6'],
#           ['к69', 'к68', 'к42', 'к41', 'к36', 'к20', 'к8' , 'п43', 'п42', 'п37', 'п15', 'п6'],
#           ['к69', 'к68', 'к37', 'к36', 'к20', 'к8' , 'п43', 'п42', 'п37', 'п15', 'п6'],
#           ['к69', 'к68', 'к37', 'к36', 'к20', 'п43', 'п42', 'п37', 'п15', 'п6'],
#           ['к69', 'к68', 'к37', 'к36', 'к20', 'п43', 'п37', 'п15', 'п6'],
#           ['к69', 'к37', 'к36', 'к20', 'п43', 'п37', 'п15', 'п6'],
#           ['к60', 'к36', 'к20', 'п43', 'п37', 'п15', 'п6'],
#           ['к60', 'к36', 'к20', 'п37', 'п15', 'п6'],
#           ['к60', 'к36', 'п37', 'п15', 'п6'],
#           ['к60', 'к36', 'п15', 'п6'],
#           ['к60', 'п15', 'п6'],
#           ['к60', 'п15'],
#           ['к60'],
#           []]


def decodPosit(lstPos, dictPos, addPos, trend):
    newPosList = []

    for posComb in lstPos:
        # print(posComb)
        newPosComb = []
        newPosList.append(dictPos[posComb][trend])

    #newPosList.append(dictPos['р6'][trend])

    #newPosList.append(dictPos['к35'][trend])
    for pos in addPos:
        newPosList.append(dictPos[pos][trend])

    newPosList.sort(reverse=True)
    # newPosList.append(newPosComb)
    return newPosList


def codPosit(lstPos, dictPos, trend):
    newPosList = []
    for posComb in sorted(lstPos):
        for k, v in dictPos.items():
            if v[trend] == posComb:
                newPosList.append(k)
    return newPosList


def calcOneComb(positionsLst, inTrend):

    returnList = []

    addPos = input("Введите доп позиции (через пробел): ").split()
    for combPosL in positionsLst:

        #combPos = decodPosit(combPosL, relevantPositions, addPos, inTrend)
        combPos = combPosL
        nKA = len(combPos)
        combPos.append(1)

        protLines = calculProtLines(combPos, matrixGraph, dictProtectLines)
        
        me = 0
        for nPos in range(nKA+1):
            me += protLines[nPos] * arrayProtectPower[combPos[nPos] - 1]

        me = me / dictProtectLines[1]
       
        txtF = open('outData.txt', 'a')
        txtF.write(str(round(me,1)) + '\n')
        txtF.close()

        #print(round(me,1), nKA, codPosit(combPos, relevantPositions, trend=curTrend))
        #returnList.append(codPosit(combPos, relevantPositions, trend=curTrend))
        print(round(me,1), nKA, combPos)
    
    #return returnList



if __name__ == '__main__':
#    
    unionPosFile = 'объедин_БелоречкаГагарский.xlsx'
    #currentCombPos = loadUnionPosit(wbName='C:/EnerTechCenter/Результаты/{}'.format(unionPosFile),
    #               sheetName='Лист1', cellsInd=('D10', 'D24'))

    #currentCombPos = [['б_17', 'б_19', 'б_28']]
    #currentCombPos = [['г_35', 'г_36', 'г_43']]
    #currentCombPos = [['б_17', 'б_19', 'б_28', 'р6', 'г_35', 'г_36', 'г_43']]

    # б_17 = 14; б_19 = 15; б_28 = 3
    # г_35 = 23; г_36 = 19; г_43 = 14
    # currentCombPos = [[15, 14, 3]] # Белоречка
    currentCombPos = [[23, 19, 14]] # Гагарский

    print('Сторону подсоединения (Тренд) не забудь!!!!!')
    curTrend = int(input('(0-Белоречка_Гагарский, 1-Гагарский_Белоречка) trend = '))

    #for i in range(4):
    #    print()
    #    calcOneComb(currentCombPos, curTrend)
    #    print(arrayProtectPower[0])
    calcOneComb(currentCombPos, curTrend)
    print(arrayProtectPower[0])
