import math
import time
from openpyxl import load_workbook
#from openpyxl import Workbook
# import xlwt
# import pandas as pd


#def writeListInXlsx(tableName, rows):
#    book = Workbook()
#    sheet = book.active
#    for row in rows:
#        # print(row)
#        sheet.append(row)
#
#    book.save('{}.xlsx'.format(tableName))

# Загрузка матрицы сопряженности линий из excel-файла
def loadMatrix(wbName, sheetName, cellsInd):
    # Load in the workbook
    wb = load_workbook(wbName)
    # Get a sheet by name
    sheet = wb.get_sheet_by_name(sheetName)

    listMatrix = []
    arrayRowLine = []
    for cellObj in sheet[cellsInd[0]:cellsInd[1]]:
        for cells in cellObj:
            arrayRowLine.append(cells.value)
        listMatrix.append(arrayRowLine.copy())
        arrayRowLine.clear()
    return listMatrix


# Загрузка данных по ТП в словарь из excel-файла в виде:
  # №_линии: (№_ТП, номин_мощн, загрузка)
def loadDataTP(wbName, sheetName, cellsInd):
    wb = load_workbook(wbName)
    sheet = wb.get_sheet_by_name(sheetName)

    dictPowerLine = {}
    for cellObj in sheet[cellsInd[0]:cellsInd[1]]:
        dictPowerLine[int(cellObj[0].value)] = (str(cellObj[1].value),
                                                int(cellObj[2].value),
                                                float(cellObj[3].value),
                                                int(cellObj[4].value))
    return dictPowerLine


 # Для каждой позиции расчитывается количество следующих за ней линий
def protectLines(inMatrix, numbPillar):
    dictProtectLines = {}
    for i in range(len(inMatrix)):
        #dictProtectLines[i+1] = sum(inMatrix[i]) #np.array(sum(inMatrix[i]))
        dictProtectLines[i+1] = sum([x * y for x, y in zip(inMatrix[i], numbPillar)])
    return dictProtectLines


# защищаемая мощность для каждой позиции КА
def functProtectConsumer(dictProtectPower, matrixGraph, amountLines):
    arrayProtectPower = []
    for posKA1 in range(0, amountLines):
        protectPower = 0
        for posKA2 in range(0, amountLines):
            if matrixGraph[posKA1][posKA2] == 1:
                listPower = dictProtectPower[posKA2 + 1]
                protectPower += listPower[1] * listPower[2]
        arrayProtectPower.append(protectPower)
    return arrayProtectPower


# проверка на наличие соседних позиций.
def notIsNeigh(listPos, dictNeigh):
    breakGenPos = True
    rowEl = listPos[-1]-1
    for el in reversed(listPos[:-1]):
        if dictNeigh[rowEl][el-1] == 1:
            breakGenPos = False
            break
    return breakGenPos


# поиск независимых позиций
def searchIndepPos(positions, matrixGraph):
    lstIndependPositions = []
    lenPos = len(positions)
    for posI in range(lenPos-1):
        isBreak = False
        for posJ in range(posI+1,lenPos):
            if matrixGraph[positions[posJ]-1][positions[posI]-1] == 1:
                isBreak = True
                break
        if isBreak: continue
        lstIndependPositions.append(positions[posI])
    lstIndependPositions.append(positions[-1])
    return tuple(lstIndependPositions)


# расчет кол-ва защищаемых линий
def calculProtLines(positions, matrixGraph, dictProtLines):
    lstProtLines = []
    dictPosAndPrL = {}
    for iPos in positions:
        # print(list(dictPosAndPrL.keys()))
        positProtLines = dictProtLines[iPos]
        #j = 0
        #lstIndPL = []
        for jPos in list(dictPosAndPrL.keys()):
        #for jPos in dictPosAndPrL:
            if matrixGraph[iPos - 1][jPos - 1] == 1:
                # if iPos == 143: print(positProtLines, jPos, dictPosAndPrL[jPos])
                positProtLines -= dictProtLines[jPos] #
                #lstIndPL.append(jPos)
                del dictPosAndPrL[jPos]
        lstProtLines.append(positProtLines)
        # for pos in lstIndPL:
        #     del dictPosAndPrL[pos]
        dictPosAndPrL[iPos] = positProtLines
    return lstProtLines

#-----------------------------------------------------------------------------------
#--------------------------------------- РАССЧЕТНАЯ ЧАСТЬ --------------------------
#-----------------------------------------------------------------------------------

feederName = input("Введите название фидера (имя папки): ")
wbName = 'C:/EnerTechCenter/ДанныеПоСетям/{}/'.format(feederName)

with open(wbName+"cellsInd.txt") as file_cells:
    for i in range(2):
        if i==0: cellDataTP = tuple(file_cells.readline().split())
        else: cellMatrix = tuple(file_cells.readline().split())

# Загрузка словаря из эксель с данными о ТП (позиция: №, ном.мощн, ст.загрузки)
dictDataTP = loadDataTP(wbName=wbName+'dataTP.xlsx', sheetName='dataTP', cellsInd=cellDataTP)


# Загрузка матрицы зависимости позиций
matrixGraph = loadMatrix(wbName=wbName+'dependMatr.xlsx', sheetName='Sheet1', cellsInd=cellMatrix)

# Формирование словаря, содержащего для каждой позиции КА количесвто зависимых позиций
dictProtectLines = protectLines(matrixGraph, [tpData[3] for tpData in dictDataTP.values()])

# Получение количества позиций на данной сети
AMOUNT_LINES = len(dictDataTP)
AMOUNT_PILLAR = dictProtectLines[1]
#print(AMOUNT_LINES)
#print(AMOUNT_PILLAR)
print(dictProtectLines)

# Расчет защищаемой мощности каждой позицией при условии, что нет ни единой КА
arrayProtectPower = functProtectConsumer(dictDataTP, matrixGraph, AMOUNT_LINES) # зщищаемая млщность каждой КА
print(arrayProtectPower[35], arrayProtectPower[0], arrayProtectPower[0])