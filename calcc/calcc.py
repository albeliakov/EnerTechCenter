import math
import time
from openpyxl import load_workbook
#from openpyxl import Workbook
# import xlwt
# import pandas as pd


#def writeListInXlsx(tableName, rows):
#    book = Workbook()
#    sheet = book.active
#    for row in rows:
#        # print(row)
#        sheet.append(row)
#
#    book.save('{}.xlsx'.format(tableName))

# Загрузка матрицы сопряженности линий из excel-файла
def loadMatrix(wbName, sheetName, cellsInd):
    # Load in the workbook
    wb = load_workbook(wbName)
    # Get a sheet by name
    sheet = wb.get_sheet_by_name(sheetName)

    listMatrix = []
    arrayRowLine = []
    for cellObj in sheet[cellsInd[0]:cellsInd[1]]:
        for cells in cellObj:
            arrayRowLine.append(cells.value)
        listMatrix.append(arrayRowLine.copy())
        arrayRowLine.clear()
    return listMatrix


# Загрузка данных по ТП в словарь из excel-файла в виде:
  # №_линии: (№_ТП, номин_мощн, загрузка)
def loadDataTP(wbName, sheetName, cellsInd):
    wb = load_workbook(wbName)
    sheet = wb.get_sheet_by_name(sheetName)

    dictPowerLine = {}
    for cellObj in sheet[cellsInd[0]:cellsInd[1]]:
        dictPowerLine[int(cellObj[0].value)] = (str(cellObj[1].value),
                                                int(cellObj[2].value),
                                                float(cellObj[3].value),
                                                int(cellObj[4].value))
    return dictPowerLine


 # Для каждой позиции расчитывается количество следующих за ней линий
def protectLines(inMatrix, numbPillar):
    dictProtectLines = {}
    for i in range(len(inMatrix)):
        #dictProtectLines[i+1] = sum(inMatrix[i]) #np.array(sum(inMatrix[i]))
        dictProtectLines[i+1] = sum([x * y for x, y in zip(inMatrix[i], numbPillar)])
    return dictProtectLines


# защищаемая мощность для каждой позиции КА
def functProtectConsumer(dictProtectPower, matrixGraph, amountLines):
    arrayProtectPower = []
    for posKA1 in range(0, amountLines):
        protectPower = 0
        for posKA2 in range(0, amountLines):
            if matrixGraph[posKA1][posKA2] == 1:
                listPower = dictProtectPower[posKA2 + 1]
                protectPower += listPower[1] * listPower[2]
        arrayProtectPower.append(protectPower)
    return arrayProtectPower


# проверка на наличие соседних позиций.
def notIsNeigh(listPos, dictNeigh):
    breakGenPos = True
    rowEl = listPos[-1]-1
    for el in reversed(listPos[:-1]):
        if dictNeigh[rowEl][el-1] == 1:
            breakGenPos = False
            break
    return breakGenPos


# поиск независимых позиций
def searchIndepPos(positions, matrixGraph):
    lstIndependPositions = []
    lenPos = len(positions)
    for posI in range(lenPos-1):
        isBreak = False
        for posJ in range(posI+1,lenPos):
            if matrixGraph[positions[posJ]-1][positions[posI]-1] == 1:
                isBreak = True
                break
        if isBreak: continue
        lstIndependPositions.append(positions[posI])
    lstIndependPositions.append(positions[-1])
    return tuple(lstIndependPositions)


# расчет кол-ва защищаемых линий
def calculProtLines(positions, matrixGraph, dictProtLines):
    lstProtLines = []
    dictPosAndPrL = {}
    for iPos in positions:
        # print(list(dictPosAndPrL.keys()))
        positProtLines = dictProtLines[iPos]
        #j = 0
        #lstIndPL = []
        for jPos in list(dictPosAndPrL.keys()):
        #for jPos in dictPosAndPrL:
            if matrixGraph[iPos - 1][jPos - 1] == 1:
                # if iPos == 143: print(positProtLines, jPos, dictPosAndPrL[jPos])
                positProtLines -= dictProtLines[jPos] #
                #lstIndPL.append(jPos)
                del dictPosAndPrL[jPos]
        lstProtLines.append(positProtLines)
        # for pos in lstIndPL:
        #     del dictPosAndPrL[pos]
        dictPosAndPrL[iPos] = positProtLines
    return lstProtLines

#-----------------------------------------------------------------------------------
#--------------------------------------- РАССЧЕТНАЯ ЧАСТЬ --------------------------
#-----------------------------------------------------------------------------------

feederName = input("Введите название фидера (имя папки): ")
wbName = 'C:/EnerTechCenter/ДанныеПоСетям/{}/'.format(feederName)

with open(wbName+"cellsInd.txt") as file_cells:
    for i in range(2):
        if i==0: cellDataTP = tuple(file_cells.readline().split())
        else: cellMatrix = tuple(file_cells.readline().split())

# Загрузка словаря из эксель с данными о ТП (позиция: №, ном.мощн, ст.загрузки)
dictDataTP = loadDataTP(wbName=wbName+'dataTP.xlsx', sheetName='dataTP', cellsInd=cellDataTP)


# Загрузка матрицы зависимости позиций
matrixGraph = loadMatrix(wbName=wbName+'dependMatr.xlsx', sheetName='Sheet1', cellsInd=cellMatrix)

# Формирование словаря, содержащего для каждой позиции КА количесвто зависимых позиций
dictProtectLines = protectLines(matrixGraph, [tpData[3] for tpData in dictDataTP.values()])
# positWithSIP = ()
# RATE_SIP = 15.9
# dictProtectLines = addWireType(dictProtectLines, positWithSIP)

# Получение количества позиций на данной сети
AMOUNT_LINES = len(dictDataTP)
AMOUNT_PILLAR = dictProtectLines[1]
print(AMOUNT_LINES)
print(AMOUNT_PILLAR)
print(dictProtectLines)

# Расчет защищаемой мощности каждой позицией при условии, что нет ни единой КА
arrayProtectPower = functProtectConsumer(dictDataTP, matrixGraph, AMOUNT_LINES) # зщищаемая млщность каждой КА

# Загрузка матрицы соседей
#sheetName = 'neighbors'
#cellsInd=('B2', 'EV152')
neighborsMatrix = loadMatrix(wbName=wbName+'matrNeigh.xlsx', sheetName='Sheet1', cellsInd=cellMatrix)

# Получение суммарной мощности на линии
SUM_POWER = arrayProtectPower[0]
# print(SUM_POWER)

# Задание изменения отключаемой мощности, при которой прекращаются расчеты (в долях)
MAX_CHANGE_POWER = 0.05

# Установка нерассчитываемых позиций
# tplLastLines = { 14,  18,  22,  23,  24,  25,  26,  27,  29,  31,  32,  33,  35,  36,  37,  39,  41,  45,  47,
#                  49,  50,  51,  53,  55,  56,  60,  62,  64,  65,  66,  68,  69,  70,  72,  76,  78,  82,  84,
#                  85,  87,  88,  89,  91,  93,  94,  98, 100, 102, 103, 105, 104, 107, 111, 113, 117, 121, 122,
#                 123, 125, 127, 129, 131, 132, 133, 134, 136, 138, 140, 142, 145, 144, 146, 147, 148, 150, 151}
tplLastLines = set()
try:
    finishValidPos = int(input("Введите № позици, после которой можно не учитывать одинарные: "))
    for pos in dictProtectLines.keys():
        if dictProtectLines[pos] == 1 and pos > finishValidPos: tplLastLines.add(pos)
except ValueError:
    print("Проссчитываются все позиции")

# for pos in dictProtectLines.keys():
#     if dictProtectLines[pos]==1 and pos>finishValidPos: tplLastLines.add(pos)
#print(tplLastLines)

def calculationFrom1ToNka(amountKAStart=1, amountKAFinish=AMOUNT_LINES-1):
    returnCalculData = [[SUM_POWER]]

    # Поиск оптимальных позиций для количества КА от 1 до amountKA
    for nKA in range(amountKAStart, amountKAFinish+1):
        txtF = open('outData.txt', 'a')
        print()
        print(nKA, 'начало:', time.ctime())
        meMin = 2*SUM_POWER #
        optimalPositions = []
        numbComb = 0

        # ИНИЦИАЛИЗАЦИЯ ПЕРВОНАЧАЛЬНЫХ ДАННЫХ
        # стартовый список позиций
        positions = []
        appPos = AMOUNT_LINES
        while appPos in tplLastLines: appPos -= 1
        positions.append(appPos)
        for i in range(nKA-1):
            appPos = positions[i] - 1
            positions.append(appPos)
            while positions[-1] != nKA-i:  # неминимальное значение
                # if lap.notIsNeigh(positions, dictNM) and not(lap.binSearch(positions[-1], tplLastLines)): break
                if notIsNeigh(positions, neighborsMatrix) and not (positions[-1] in tplLastLines): break
                positions[-1] -= 1

        if nKA > 1:
            # кортеж, содержащий независмые позиции, исключая последнюю позицию (для расчета для 1 и новой позиций)
            tplIndPosWithoutLast = searchIndepPos(positions[:-1], matrixGraph)

            # список, содержащий кол-во защищаемых линий, исключая для последней позиции
            lstProtLinesWithoutLast = calculProtLines(positions[:-1], matrixGraph, dictProtectLines)

        else:
            tplIndPosWithoutLast = ()
            lstProtLinesWithoutLast = []

        # ЦИКЛ ФОРМИРОВАНИЯ КОМБИНАЦИЙ И ИХ ПРОСЧЕТА
        while True:
            # print(positions)
            numbComb += 1
            if numbComb%100000000==0: print('Кол-во просчитанных комбинаций: ', numbComb)
            lastPosit = positions[-1] # последняя в списке позиция

            # ФОрмирование списка независимых позиций и списка с кол-вом защищаемых линий для каждой позиции
            lstIndPos = []
            lstProtLines = lstProtLinesWithoutLast.copy()
            lastProtLines = dictProtectLines[lastPosit]
            for indPos in tplIndPosWithoutLast:
                if matrixGraph[lastPosit-1][indPos-1] == 0:
                    lstIndPos.append(indPos)
                else:
                    lastProtLines -= dictProtectLines[indPos]
            lstIndPos.append(lastPosit)
            lstProtLines.append(lastProtLines)

            # Рассчет кол-ва защишаемых линий позицией 1
            oneProtLines = dictProtectLines[1]
            for indPos in lstIndPos:
                oneProtLines -= dictProtectLines[indPos]

            # Рассчет мат ожидания
            me = 0
            for nPos in range(nKA):
                me += lstProtLines[nPos] * arrayProtectPower[positions[nPos] - 1]
            me = (me + oneProtLines * arrayProtectPower[0]) / AMOUNT_PILLAR
            if me < meMin:
                optimalPositions.clear()
                meMin = me
                optimalPositions.append(positions.copy())
            elif me == meMin:
                optimalPositions.append(positions.copy())  # если оптимальных позиций окажется несколько

            # Условие оканчание расчетов для соответсвующего nKA
            if positions[0] == nKA + 1: break

            indBreak = 0
            for i in range(-1, -(nKA + 1), -1):
                if positions[i] != -i + 1:
                    positions[i] = positions[i] - 1
                    while positions[i] != -i + 1: # неминимальное значение
                        # if lap.notIsNeigh(positions[:len(positions)+i+1], dictNM) and not(lap.binSearch(positions[i], tplLastLines)): break
                        if notIsNeigh(positions[:nKA + i + 1], neighborsMatrix) and not (positions[i] in tplLastLines): break
                        positions[i] = positions[i] - 1

                    for ind in range(i+1, 0, 1):
                        positions[ind] = positions[ind - 1] - 1
                        while positions[ind] != -ind + 1:
                            # if lap.notIsNeigh(positions[:len(positions)+ind+1], dictNM) and not(lap.binSearch(positions[ind], tplLastLines)): break
                            if notIsNeigh(positions[:nKA + ind + 1], neighborsMatrix) and not (positions[ind] in tplLastLines): break
                            positions[ind] = positions[ind] - 1
                    indBreak = i
                    break
            if indBreak < -1:
                # формирование новых списков независимых позиций и кол-ва защищаемых линий
                tplIndPosWithoutLast = searchIndepPos(positions[:-1], matrixGraph)
                lstProtLinesWithoutLast = calculProtLines(positions[:-1], matrixGraph, dictProtectLines)

        # Приостановка рассчетов при достижения изменения отключаемой мощности не болле, чем на XX%
        # if 1 - (meMin / returnCalculData[-1][0]) <= MAX_CHANGE_POWER:
        #     returnCalculData.append((round(meMin, 1), nKA, optimalPositions.copy()))
        #     print('ВНИМАНИЕ: Достигли изменения отключаемой мощности не более, чем на {:.1%}'.format(MAX_CHANGE_POWER))
        #     print('          при количестве установленных КА = {}'.format(nKA))
        #     break

        txtF.write(str((round(meMin, 1), nKA, optimalPositions.copy())) + '\n')
        txtF.close()
        print((round(meMin, 1), nKA, optimalPositions.copy()))
        returnCalculData.append((round(meMin, 1), nKA, optimalPositions.copy()))
    print('\nКонец рассчетов:', time.ctime())

    return returnCalculData[1:]


dictCurrentPos = {'к8'  : [8, 133],  'к20' : [20, 102], 'к21' : [21, 97], 'к35' : [35, 83],
                  'к36' : [36, 82],  'к37' : [37, 84],  'к41' : [41, 87], 'к42' : [42, 88], 
                  'к48' : [48, 94], 'к60' : [60, 106],  'к68' : [72, 126], 'к69' : [73, 125], 
                  'п6'  : [146, 6], 'п8'  : [144, 8],   'п15' : [141, 15], 'п37' : [93, 37],  
                  'п42' : [81, 42], 'п43' : [88, 43],   'шр'  : [67, 62]}

posList = [['к69', 'к68', 'к42', 'к41', 'к36', 'к21', 'к20', 'к8' , 'п43', 'п42', 'п37', 'п15', 'п8', 'п6'],
           ['к69', 'к68', 'к42', 'к41', 'к36', 'к21', 'к20', 'к8' , 'п43', 'п42', 'п37', 'п15', 'п6'],
           ['к69', 'к68', 'к42', 'к41', 'к36', 'к20', 'к8' , 'п43', 'п42', 'п37', 'п15', 'п6'],
           ['к69', 'к68', 'к37', 'к36', 'к20', 'к8' , 'п43', 'п42', 'п37', 'п15', 'п6'],
           ['к69', 'к68', 'к37', 'к36', 'к20', 'п43', 'п42', 'п37', 'п15', 'п6'],
           ['к69', 'к68', 'к37', 'к36', 'к20', 'п43', 'п37', 'п15', 'п6'],
           ['к69', 'к37', 'к36', 'к20', 'п43', 'п37', 'п15', 'п6'],
           ['к60', 'к36', 'к20', 'п43', 'п37', 'п15', 'п6'],
           ['к60', 'к36', 'к20', 'п37', 'п15', 'п6'],
           ['к60', 'к36', 'п37', 'п15', 'п6'],
           ['к60', 'к36', 'п15', 'п6'],
           ['к60', 'п15', 'п6'],
           ['к60', 'п15'],
           ['к60'],
           []]

# сделать листы с соответсвующими номерами, добавив в каждый 'шр' и сортировка по убыванию
# Кашино_ППЗ - 0
# ППЗ_Кашино - 1
#curTrend = 0
print('Тренд не забудь!!!!!')
curTrend = int(input('(0-Кашино_ППЗ, 1-ППЗ_Кашино) trend = '))
def decodPosit(lstPos, dictPos, trend):
    newPosList = []
    for posComb in lstPos:
        # print(posComb)
        newPosComb = []
        newPosList.append(dictPos[posComb][trend])
    newPosList.append(dictPos['шр'][trend])
    newPosList.append(dictPos['к35'][trend])
    newPosList.sort(reverse=True)
    # newPosList.append(newPosComb)
    return newPosList

def codPosit(lstPos, dictPos, trend):
    newPosList = []
    for posComb in sorted(lstPos):
        for k, v in dictPos.items():
            if v[trend] == posComb:
                newPosList.append(k)
    return newPosList


def calcOneComb(positionsLst, inTrend):
    
    returnList = []
    for combPosL in positionsLst:
        # print(combPosL)
        combPos = decodPosit(combPosL, dictCurrentPos, inTrend)
        nKA = len(combPos)

        if nKA > 1:
            # кортеж, содержащий независмые позиции, исключая последнюю позицию (для расчета для 1 и новой позиций)
            tplIndPosWithoutLast = searchIndepPos(combPos[:-1], matrixGraph)
            # список, содержащий кол-во защищаемых линий, исключая для последней позиции
            lstProtLinesWithoutLast = calculProtLines(combPos[:-1], matrixGraph, dictProtectLines)
        else:
            tplIndPosWithoutLast = ()
            lstProtLinesWithoutLast = []

        # ЦИКЛ ФОРМИРОВАНИЯ КОМБИНАЦИЙ И ИХ ПРОСЧЕТА
        lastPosit = combPos[-1] # последняя в списке позиция

        # ФОрмирование списка независимых позиций и списка с кол-вом защищаемых линий для каждой позиции
        lstIndPos = []
        lstProtLines = lstProtLinesWithoutLast.copy()
        lastProtLines = dictProtectLines[lastPosit]
        for indPos in tplIndPosWithoutLast:
            if matrixGraph[lastPosit-1][indPos-1] == 0:
                lstIndPos.append(indPos)
            else:
                lastProtLines -= dictProtectLines[indPos]
        lstIndPos.append(lastPosit)
        lstProtLines.append(lastProtLines)

        # Рассчет кол-ва защишаемых линий позицией 1
        oneProtLines = dictProtectLines[1]
        for indPos in lstIndPos:
            oneProtLines -= dictProtectLines[indPos]

        # Рассчет мат ожидания
        me = 0
        for nPos in range(nKA):
            me += lstProtLines[nPos] * arrayProtectPower[combPos[nPos] - 1]
        me = (me + oneProtLines * arrayProtectPower[0]) / AMOUNT_PILLAR

        # print(combPos)
        print(round(me,1), nKA, codPosit(combPos, dictCurrentPos, trend=curTrend))
        returnList.append(codPosit(combPos, dictCurrentPos, trend=curTrend))
    # print(returnList)
    return returnList



def calculationLastNka(kaLast, amountKA = AMOUNT_LINES-1):
    returnCalculData = [[SUM_POWER]]

    # Поиск оптимальных позиций для количества КА от 1 до amountKA
    for nKA in range(amountKA-kaLast+1, amountKA+1):
        print()
        print(nKA, 'начало:', time.ctime())
        meMin = 2*SUM_POWER #
        optimalPositions = []
        numbComb = 0

        # ИНИЦИАЛИЗАЦИЯ ПЕРВОНАЧАЛЬНЫХ ДАННЫХ
        # стартовый список позиций
        # positions = [AMOUNT_LINES]
        positions = [i for i in range(AMOUNT_LINES, AMOUNT_LINES-nKA, -1)]

        if nKA > 1:
            # кортеж, содержащиq независмые позиции, исключая последнюю позицию (для расчета для 1 и новой позиций)
            tplIndPosWithoutLast = searchIndepPos(positions[:-1], matrixGraph)
            # список, содержащий кол-во защищаемых линий, исключая для последней позиции
            lstProtLinesWithoutLast = calculProtLines(positions[:-1], matrixGraph, dictProtectLines)
        else:
            tplIndPosWithoutLast = ()
            lstProtLinesWithoutLast = []

        # ЦИКЛ ФОРМИРОВАНИЯ КОМБИНАЦИЙ И ИХ ПРОСЧЕТА
        while True:
            # print(positions)
            numbComb += 1
            if numbComb%1000000==0: print('Кол-во просчитанных комбинаций: ', numbComb)
            lastPosit = positions[-1] # последняя в списке позиция

            # ФОрмирование списка независимых позиций и списка с кол-вом защищаемых линий для каждой позиции
            lstIndPos = []
            lstProtLines = lstProtLinesWithoutLast.copy()
            lastProtLines = dictProtectLines[lastPosit]
            for indPos in tplIndPosWithoutLast:
                if matrixGraph[lastPosit-1][indPos-1] == 0:
                    lstIndPos.append(indPos)
                else:
                    lastProtLines -= dictProtectLines[indPos]
            lstIndPos.append(lastPosit)
            lstProtLines.append(lastProtLines)

            # Рассчет кол-ва защишаемых линий позицией 1
            oneProtLines = dictProtectLines[1]
            for indPos in lstIndPos:
                oneProtLines -= dictProtectLines[indPos]
              #print(lstProtLines, oneProtLines)
            # Рассчет мат ожидания
            me = 0
            for nPos in range(nKA):
                me += lstProtLines[nPos] * arrayProtectPower[positions[nPos] - 1]
            me = (me + oneProtLines * arrayProtectPower[0]) / AMOUNT_PILLAR
            if me < meMin:
                optimalPositions.clear()
                meMin = me
                optimalPositions.append(positions.copy())
            elif me == meMin:
                optimalPositions.append(positions.copy())  # если оптимальных позиций окажется несколько

            # Условие оканчание расчетов для соответсвующего nKA
            if positions[0] == nKA + 1: break

            indBreak = 0
            for i in range(-1, -(nKA + 1), -1):
                if positions[i] != -i + 1:
                    positions[i] = positions[i] - 1

                    for ind in range(i+1, 0, 1):
                        positions[ind] = positions[ind - 1] - 1

                    indBreak = i
                    break
            if indBreak < -1:
                # формирование новых списков независимых позиций и кол-ва защищаемых линий
                tplIndPosWithoutLast = searchIndepPos(positions[:-1], matrixGraph)
                lstProtLinesWithoutLast = calculProtLines(positions[:-1], matrixGraph, dictProtectLines)

        print((round(meMin, 1), nKA, optimalPositions.copy()))
        returnCalculData.append((round(meMin, 1), nKA, optimalPositions.copy()))
    print('\nКонец рассчетов:', time.ctime())

    return returnCalculData[1:]



# Задание количества КА для проведения рассчетов
#kaFirst, kaLast = map(int, input().split())
print("Введите количесвто КА,")
kaFirstStart = int(input("с которого начинать рассчет: "))
kaFirstFinish = int(input("которым закончить: "))

kaLast = int(input("\nВведите число значений количесвта КА, рассчитываемых в конце: "))


calculationFrom1ToNka(kaFirstStart, kaFirstFinish)
# calculationLastNka(kaLast)

# calcOneComb(positionsLst=posList, inTrend=curTrend)
#writeListInXlsx('Кашино_ППЗ', calcOneComb(positionsLst=posList, inTrend=curTrend)) #<-----
#writeGrMatrInXlsx('ППЗ_Кашино', calcOneComb(positionsLst=posList, inTrend=1))